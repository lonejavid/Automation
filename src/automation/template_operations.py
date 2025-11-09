"""
Template Operations Module
Handles pasting data, formulas, pivot tables, dates
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import shutil


def create_backup(template_path):
    """Create backup of template"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = template_path.replace('.xlsx', f'_backup_{timestamp}.xlsx')
    shutil.copy2(template_path, backup_path)
    print(f"\n✅ Backup created: {backup_path}")
    return backup_path


def paste_to_template(data_frames, template_path, target_sheet='AllStores'):
    """
    Paste transformed data into template
    
    Args:
        data_frames: List of DataFrames (one per store)
        template_path: Path to Drive Thru template
        target_sheet: Sheet name to paste into
    
    Returns:
        Workbook object
    """
    print(f"\n📋 Pasting data into template...")
    
    # Load template
    wb = load_workbook(template_path)
    
    # Find target sheet
    if target_sheet not in wb.sheetnames:
        print(f"   ⚠️  '{target_sheet}' not found. Available: {wb.sheetnames}")
        # Try alternatives
        for alt in ['AllStores', 'Allstores', 'Raw Data', 'RawData']:
            if alt in wb.sheetnames:
                target_sheet = alt
                print(f"   Using '{target_sheet}' instead")
                break
    
    ws = wb[target_sheet]
    
    # Find last row with data
    last_row = ws.max_row
    start_row = last_row + 1
    
    print(f"   Target sheet: {target_sheet}")
    print(f"   Current last row: {last_row}")
    print(f"   Pasting from row: {start_row}")
    
    # Combine all data frames
    combined_df = pd.concat(data_frames, ignore_index=True)
    
    # Paste data (without headers)
    row_count = 0
    for r_idx, row in enumerate(dataframe_to_rows(combined_df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        row_count += 1
    
    print(f"   ✅ Pasted {row_count} rows to '{target_sheet}'")
    
    return wb


def concatenate_formulas(wb, sheet_name, start_row, end_row, formula_columns):
    """
    Copy formulas down for yellow-headed columns
    
    Args:
        wb: Workbook object
        sheet_name: Sheet name
        start_row: First row to copy formula to
        end_row: Last row to copy formula to
        formula_columns: List of column indices with formulas (e.g., [12, 13, 14])
    """
    print(f"\n🔄 Concatenating formulas in '{sheet_name}'...")
    
    ws = wb[sheet_name]
    
    for col_idx in formula_columns:
        # Get formula from row before start_row
        source_formula = ws.cell(row=start_row - 1, column=col_idx).value
        
        if source_formula and isinstance(source_formula, str) and source_formula.startswith('='):
            # Copy formula down
            for row_idx in range(start_row, end_row + 1):
                ws.cell(row=row_idx, column=col_idx).value = source_formula
            
            print(f"   ✅ Column {col_idx}: Formula copied down {end_row - start_row + 1} rows")
    
    return wb


def refresh_pivot_tables(wb):
    """
    Set pivot tables to refresh on open
    
    Args:
        wb: Workbook object
    
    Returns:
        Workbook object
    """
    print(f"\n🔄 Setting pivot tables to refresh on open...")
    
    pivot_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if hasattr(ws, '_pivots') and ws._pivots:
            for pivot in ws._pivots:
                pivot.cache.refreshOnLoad = True
                pivot_count += 1
    
    print(f"   ✅ Set {pivot_count} pivot tables to auto-refresh")
    print(f"   Note: Full refresh happens when you open the file in Excel")
    
    return wb


def update_dates(wb, target_date, sheet_configs):
    """
    Update dates in specified sheets
    
    Args:
        wb: Workbook object
        target_date: Date to set
        sheet_configs: Dict of {sheet_name: cell_reference}
    
    Returns:
        Workbook object
    """
    print(f"\n📅 Updating dates to: {target_date.strftime('%Y-%m-%d')}")
    
    for sheet_name, cell_ref in sheet_configs.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws[cell_ref] = target_date
            print(f"   ✅ {sheet_name}[{cell_ref}] = {target_date.strftime('%Y-%m-%d')}")
        else:
            print(f"   ⚠️  Sheet '{sheet_name}' not found")
    
    return wb


def save_template(wb, output_path):
    """Save the workbook"""
    print(f"\n💾 Saving template...")
    wb.save(output_path)
    wb.close()
    print(f"   ✅ Saved: {output_path}")



