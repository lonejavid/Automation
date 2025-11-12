"""
Run VBA Macro on Downloaded Excel File
Executes the DT macro (Ctrl+D) on the downloaded Raw Car Data report
"""

import os
import glob
import sys
import subprocess
import time
from pathlib import Path
from datetime import datetime
from collections import defaultdict

DT_MACRO_CODE = """Sub DT()
'
' DT Macro
'
' Keyboard Shortcut: Ctrl+d
'
    Cells.Select
    Selection.UnMerge
    Range("B7").Select
    ActiveCell.FormulaR1C1 = "Store Name"
    Range("B4").Select
    Selection.Copy
    Range("B8").Select
    ActiveSheet.Paste
    Rows("1:6").Select
    Range("A6").Activate
    Application.CutCopyMode = False
    Selection.Delete Shift:=xlUp
    Range("B2").Select
    Selection.AutoFill Destination:=Range("B2:B197")
    Range("B2:B197").Select
    Columns("D:D").Select
    Selection.Delete Shift:=xlToLeft
    Columns("E:F").Select
    Selection.Delete Shift:=xlToLeft
    Columns("G:G").Select
    Selection.Delete Shift:=xlToLeft
    Columns("I:I").ColumnWidth = 1.57
    Columns("I:K").Select
    Selection.Delete Shift:=xlToLeft
    Columns("K:K").Select
    Selection.Delete Shift:=xlToLeft
    Columns("J:J").ColumnWidth = 4
    Columns("J:J").Select
    Selection.Delete Shift:=xlToLeft
    Columns("B:B").ColumnWidth = 19.86
    Columns("B:B").ColumnWidth = 33.29
    Range("B6").Select
    Range(Selection, Selection.End(xlDown)).Select
    Range("A197").Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range(Selection, Selection.End(xlUp)).Select
    Range("A2:A197").Select
    Range("A197").Activate
    Selection.SpecialCells(xlCellTypeBlanks).Select
    Selection.FormulaR1C1 = "=R[-1]C"
    Columns("F:G").Delete Shift:=xlToLeft
    Columns("J:K").Delete Shift:=xlToLeft
End Sub
"""

def open_excel_file(excel_file_path):
    """Open the Excel workbook in the system's default spreadsheet application."""
    try:
        if sys.platform.startswith("darwin"):  # macOS
            subprocess.Popen(["open", excel_file_path])
        elif os.name == "nt":  # Windows
            os.startfile(excel_file_path)  # type: ignore[attr-defined]
        else:  # Linux / other
            subprocess.Popen(["xdg-open", excel_file_path])
        print(f"   📂 Excel opened: {excel_file_path}")
    except Exception as e:
        print(f"   ⚠️  Unable to open Excel automatically: {e}")


def wait_for_file_download(file_path, max_wait=60):
    """
    Wait for a file to be fully downloaded (not still being written)
    
    Args:
        file_path: Path to the file
        max_wait: Maximum seconds to wait
    
    Returns:
        True if file is ready, False if timeout
    """
    start_time = time.time()
    last_size = -1
    
    while time.time() - start_time < max_wait:
        if os.path.exists(file_path):
            try:
                current_size = os.path.getsize(file_path)
                # If file size hasn't changed in 2 seconds, it's done downloading
                if current_size == last_size and current_size > 0:
                    time.sleep(1)  # One more second to be sure
                    return True
                last_size = current_size
            except (OSError, IOError):
                # File might be locked, wait a bit
                pass
        time.sleep(0.5)
    
    return os.path.exists(file_path) and os.path.getsize(file_path) > 0


def unblock_downloaded_file(file_path):
    """
    Unblock a downloaded file on Windows (removes Zone.Identifier)
    This prevents Excel from opening in Protected View
    
    Args:
        file_path: Path to the file
    """
    if sys.platform == "win32":
        try:
            zone_file = file_path + ":Zone.Identifier"
            if os.path.exists(zone_file):
                os.remove(zone_file)
                print(f"   🔓 Unblocked Windows file: {os.path.basename(file_path)}")
        except Exception as e:
            print(f"   ⚠️  Could not unblock file: {e}")


def find_latest_downloaded_file(downloads_folder, wait_for_download=True):
    """
    Find the most recently downloaded Excel file
    
    Args:
        downloads_folder: Path to downloads directory
        wait_for_download: If True, wait for file to finish downloading
    
    Returns:
        Path to latest Excel file or None
    """
    excel_files = glob.glob(os.path.join(downloads_folder, "*.xlsx"))
    
    if not excel_files:
        return None
    
    # Sort by modification time, most recent first
    latest_file = max(excel_files, key=os.path.getmtime)
    
    if wait_for_download:
        print(f"   ⏳ Waiting for file to finish downloading...")
        if wait_for_file_download(latest_file, max_wait=60):
            print(f"   ✅ File download complete")
        else:
            print(f"   ⚠️  File may still be downloading, proceeding anyway...")
    
    # Unblock file on Windows
    unblock_downloaded_file(latest_file)
    
    return latest_file


def inject_dt_macro(wb):
    """
    Inject the DT macro into the workbook if it is missing.

    Requires Excel setting "Trust access to the VBA project object model".

    Args:
        wb: xlwings workbook object

    Returns:
        True if macro injected successfully, False otherwise.
    """
    try:
        vbproject = wb.api.VBProject
    except Exception as e:
        print("   ❌ Excel blocked access to the VBA project.")
        print("   💡 Enable 'Trust access to the VBA project object model' in Trust Center.")
        print(f"   System message: {e}")
        return False

    try:
        # Check if a module named ModuleDT already exists
        module = None
        for component in vbproject.VBComponents:
            if component.Name.lower() == "moduledt" or component.Name.lower() == "dtmodule":
                module = component
                break

        if module is None:
            module = vbproject.VBComponents.Add(1)  # 1 = vbext_ct_StdModule
            module.Name = "ModuleDT"

        code_module = module.CodeModule
        # Clear existing code
        existing_lines = code_module.CountOfLines
        if existing_lines > 0:
            code_module.DeleteLines(1, existing_lines)

        # Add DT macro code
        code_module.AddFromString(DT_MACRO_CODE)
        print("   ✅ DT macro injected into workbook")
        return True
    except Exception as e:
        print(f"   ❌ Failed to inject DT macro: {e}")
        return False


def ensure_excel_edit_mode(wb):
    """
    Ensure Excel workbook is in edit mode (not Protected View)
    
    Args:
        wb: xlwings workbook object
    """
    try:
        if wb.api.ProtectStructure or wb.api.ProtectWindows:
            print("   ⚠️  Workbook is protected, attempting to enable editing...")
            # Try to enable editing
            wb.api.Application.DisplayAlerts = False
            wb.api.Application.EnableEvents = True
    except:
        pass


def run_dt_macro_excel_actions(excel_file_path):
    """
    Execute the DT macro steps using Excel automation via xlwings.
    This does not require the DT VBA macro to be installed.
    """
    try:
        import xlwings as xw
    except ImportError:
        print("   ⚠️  xlwings not installed for Excel actions fallback.")
        return False

    print("   🔄 Running DT transformations via Excel automation...")

    app = xw.App(visible=True)
    app.api.DisplayAlerts = False
    app.api.ScreenUpdating = True
    success = False

    try:
        wb = app.books.open(excel_file_path, read_only=False)
        sht = wb.sheets[0]

        # Step 1: Unmerge all cells
        sht.cells.unmerge()

        # Step 2: Prepare store name and header
        store_name = sht["B4"].value
        sht["B7"].value = "Store Name"

        last_row = sht.range("A" + str(sht.cells.last_cell.row)).end("up").row
        if store_name:
            sht.range(f"B8:B{last_row}").value = store_name

        # Step 3: Delete top rows
        sht.range("1:6").api.Delete(Shift=-4159)  # xlShiftToLeft

        # Step 4: Ensure column B is filled with store name
        last_row = sht.range("B" + str(sht.cells.last_cell.row)).end("up").row
        if store_name:
            sht.range(f"B2:B{last_row}").value = store_name

        # Step 5: Column deletions and adjustments (exact order from macro)
        sht.range("D:D").api.Delete(Shift=-4159)
        sht.range("E:F").api.Delete(Shift=-4159)
        sht.range("G:G").api.Delete(Shift=-4159)
        sht.range("I:I").column_width = 1.57
        sht.range("I:K").api.Delete(Shift=-4159)
        sht.range("K:K").api.Delete(Shift=-4159)
        sht.range("J:J").column_width = 4
        sht.range("J:J").api.Delete(Shift=-4159)
        sht.range("B:B").column_width = 19.86
        sht.range("B:B").column_width = 33.29

        # Step 6: Fill blank dayparts using the value above
        last_row = sht.range("A" + str(sht.cells.last_cell.row)).end("up").row
        try:
            sht.range(f"A2:A{last_row}").api.SpecialCells(4).FormulaR1C1 = "=R[-1]C"  # xlCellTypeBlanks = 4
        except Exception:
            pass

        # Final column removals
        sht.range("F:G").api.Delete(Shift=-4159)
        sht.range("J:K").api.Delete(Shift=-4159)

        # Save changes (Excel stays open for the user)
        wb.save()
        print("   ✅ Excel automation completed successfully")
        print("   📂 Excel workbook left open for review")
        print("\n" + "="*80)
        print("✅ FILE CONVERTED SUCCESSFULLY!")
        print("="*80)
        print(f"   📁 File: {os.path.basename(excel_file_path)}")
        print(f"   📍 Location: {excel_file_path}")
        print("="*80)
        success = True
        return True
    except Exception as e:
        print(f"   ❌ Excel automation fallback failed: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        if not success:
            try:
                for wb in app.books:
                    wb.close()
            except Exception:
                pass
            try:
                app.quit()
            except Exception:
                pass


def run_dt_macro(excel_file_path, app=None):
    """
    Run the DT macro on the Excel file using xlwings
    
    Args:
        excel_file_path: Path to Excel file
    
    Returns:
        True if successful, False otherwise
    """
    app_provided = app is not None
    wb = None

    try:
        import xlwings as xw
        
        print(f"\n📝 Opening Excel file: {os.path.basename(excel_file_path)}")
        
        if not app_provided:
            app = xw.App(visible=True)
        try:
            app.visible = True
        except Exception:
            pass
        
        try:
            app.api.DisplayAlerts = False
            app.api.ScreenUpdating = True
        except Exception:
            pass
        
        # Open workbook
        try:
            wb = app.books.open(excel_file_path, read_only=False)
        except Exception as e:
            print(f"   ⚠️  Could not open file with xlwings: {e}")
            print("   🔄 Falling back to Python logic...")
            if not app_provided:
                try:
                    app.quit()
                except Exception:
                    pass
            return run_dt_macro_python_logic(excel_file_path)
        
        # Ensure edit mode
        ensure_excel_edit_mode(wb)
        
        print("   Running DT macro (Ctrl+D equivalent)...")
        
        # Try multiple methods to run the macro
        macro_success = False
        
        # Method 1: Try to run from add-in or workbook
        try:
            app.api.Run("DT")
            print("   ✅ Macro executed successfully via DT")
            macro_success = True
        except Exception as macro_error:
            # Method 2: Try with full path
            try:
                app.api.Run(f"'{os.path.basename(excel_file_path)}'!DT")
                print("   ✅ Macro executed successfully (with workbook name)")
                macro_success = True
            except:
                # Method 3: Try from Personal.xlsb or add-in
                try:
                    app.api.Run("Personal.xlsb!DT")
                    print("   ✅ Macro executed successfully from Personal.xlsb")
                    macro_success = True
                except:
                    pass
        
        if not macro_success:
            print("   ⚠️  Macro 'DT' not found in add-ins or workbook.")
            print("   🔄 Attempting to inject DT macro into workbook...")
            if inject_dt_macro(wb):
                try:
                    app.api.Run(f"'{wb.name}'!DT")
                    print("   ✅ Macro executed successfully after injection")
                    macro_success = True
                except Exception as injected_run_error:
                    print(f"   ❌ Could not run injected macro: {injected_run_error}")
            else:
                print("   ⚠️  Injection failed or blocked by Excel security.")

        if macro_success:
            try:
                wb.save()
                print("   💾 File saved successfully")
            except Exception as save_error:
                print(f"   ⚠️  Could not save automatically: {save_error}")
                print("   📝 File is open in Excel - please save manually if needed")
            
            try:
                wb.activate()
            except Exception:
                try:
                    wb.api.Activate()
                except Exception:
                    pass
            
            print("   📂 Excel workbook left open for review")
            
            print("\n" + "="*80)
            print("✅ FILE CONVERTED SUCCESSFULLY!")
            print("="*80)
            print(f"   📁 File: {os.path.basename(excel_file_path)}")
            print(f"   📍 Location: {excel_file_path}")
            print("="*80)
            return True
        else:
            print("   ⚠️  DT macro could not be executed via Excel.")
            print("   🔄 Trying Excel automation sequence as fallback...")
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
            if not app_provided:
                try:
                    app.quit()
                except Exception:
                    pass
            excel_actions_success = run_dt_macro_excel_actions(excel_file_path)
            if excel_actions_success:
                return True
            print("   🔄 Replicating macro logic in Python...")
            try:
                if wb is not None:
                    wb.close()
            except Exception:
                pass
            if not app_provided:
                try:
                    app.quit()
                except Exception:
                    pass
            return run_dt_macro_python_logic(excel_file_path)
    except ImportError:
        print("   ⚠️  xlwings not installed. Replicating macro logic in Python...")
        return run_dt_macro_python_logic(excel_file_path)
    except Exception as e:
        print(f"   ❌ Error running macro with xlwings: {e}")
        import traceback
        traceback.print_exc()
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
        if app is not None and not app_provided:
            try:
                app.quit()
            except Exception:
                pass
        print("   🔄 Falling back to Python logic...")
        return run_dt_macro_python_logic(excel_file_path)


def run_dt_macro_python_logic(excel_file_path, wb=None, ws=None):
    """
    Replicate the DT macro logic using openpyxl
    This is a fallback if xlwings is not available or macro doesn't exist
    """
    should_close = False
    try:
        from openpyxl import load_workbook
        
        if wb is None:
            print(f"\n📝 Opening Excel file: {os.path.basename(excel_file_path)}")
            wb = load_workbook(excel_file_path)
            ws = wb.active
            should_close = True
            try:
                headers = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
                print(f"   🔍 Raw headers (row 1): {headers}")
                for debug_row in range(1, 8):
                    preview = [
                        ws.cell(row=debug_row, column=col_idx).value
                        for col_idx in range(1, min(ws.max_column, 12) + 1)
                    ]
                    print(f"   🔹 Raw row {debug_row}: {preview}")
            except Exception:
                pass
        
        print("   Replicating DT macro logic...")
        
        # Step 1: Unmerge all cells
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            ws.unmerge_cells(str(merged_range))
        
        # Step 2: Add "Store Name" header at B7
        ws['B7'] = "Store Name"
        
        # Step 3: Copy store name from B4 to column B starting at B8
        store_name = ws['B4'].value
        if store_name:
            max_row = ws.max_row
            for row in range(8, min(max_row + 1, 198)):  # B8:B197
                ws[f'B{row}'] = store_name
        
        # Step 4: Delete rows 1-6
        ws.delete_rows(1, 6)
        try:
            headers_after_row_delete = [ws.cell(row=1, column=idx).value for idx in range(1, ws.max_column + 1)]
            print(f"   🔍 Headers after row delete: {headers_after_row_delete}")
        except Exception:
            pass
        
        # After deleting rows 1-6, ensure Store Name and Daypart columns are filled down
        max_row = ws.max_row
        if store_name:
            for row in range(2, max_row + 1):
                cell = ws.cell(row=row, column=2)
                if cell.value in (None, ""):
                    cell.value = store_name
        
        for row in range(2, max_row + 1):
            daypart_cell = ws.cell(row=row, column=1)
            if daypart_cell.value in (None, ""):
                daypart_cell.value = ws.cell(row=row - 1, column=1).value
        
        # Build header map to locate columns by header text
        header_map = {}
        for col_idx in range(1, ws.max_column + 1):
            header_value = ws.cell(row=1, column=col_idx).value
            if header_value is None:
                continue
            header_key = str(header_value).strip()
            if header_key not in header_map:
                header_map[header_key] = []
            header_map[header_key].append(col_idx)
        
        desired_columns = [
            ("Daypart", 1),
            ("Store Name", 2),
            ("Departure Time", 3),
            ("Event Name", 5),
            ("Cars in Queue", 8),
            ("Cars In Order Queue", 9),
            ("Cars In Order Point Stack", 11),
            ("Menu Board", 12),
            ("Greet", 16),
            ("Service", 19),
            ("Lane Queue", 20),
            ("Lane Total", 23),
            ("Lane Total 2", 24),
        ]
        
        original_title = ws.title
        formatted_ws = wb.create_sheet(title=f"{original_title}_formatted")
        
        for new_col_idx, (header, fallback_index) in enumerate(desired_columns, start=1):
            source_col = None
            if header in header_map and header_map[header]:
                source_col = header_map[header].pop(0)
            else:
                source_col = fallback_index
            
            formatted_ws.cell(row=1, column=new_col_idx).value = header
            for row_idx in range(2, max_row + 1):
                formatted_ws.cell(row=row_idx, column=new_col_idx).value = ws.cell(
                    row=row_idx, column=source_col
                ).value
        
        # Replace old worksheet with new formatted version
        wb.remove(ws)
        formatted_ws.title = original_title
        ws = formatted_ws
        
        # Set key column widths similar to macro output
        ws.column_dimensions['B'].width = 33.29
        
        # Save changes
        wb.save(excel_file_path)
        if should_close and wb:
            wb.close()
        
        print("   ✅ Macro logic replicated and applied")
        print("\n" + "="*80)
        print("✅ FILE CONVERTED SUCCESSFULLY!")
        print("="*80)
        print(f"   📁 File: {os.path.basename(excel_file_path)}")
        print(f"   📍 Location: {excel_file_path}")
        print("="*80)
        
        # Attempt to open the workbook in Excel for the user
        open_excel_file(excel_file_path)
        return True
        
    except Exception as e:
        print(f"   ❌ Error replicating macro: {e}")
        import traceback
        traceback.print_exc()
        if should_close and wb:
            try:
                wb.close()
            except:
                pass
        return False


def process_downloaded_file(downloads_folder=None, wait_for_download=True):
    """
    Find latest downloaded file and run DT macro on it automatically
    
    This function is called automatically after each download to ensure
    the DT macro runs on every downloaded Excel file.
    
    Args:
        downloads_folder: Path to downloads folder (defaults to data/downloads)
        wait_for_download: If True, wait for file to finish downloading
    
    Returns:
        True if successful, False otherwise
    """
    if downloads_folder is None:
        BASE_DIR = Path(__file__).resolve().parents[2]
        downloads_folder = BASE_DIR / "data" / "downloads"
    
    downloads_folder = str(downloads_folder)
    
    print("\n" + "="*80)
    print("🔄 RUNNING DT MACRO ON DOWNLOADED FILE")
    print("="*80)
    
    # Ensure downloads folder exists
    if not os.path.exists(downloads_folder):
        print(f"   ❌ Downloads folder does not exist: {downloads_folder}")
        return False
    
    # Find latest file (and wait for download to complete)
    latest_file = find_latest_downloaded_file(downloads_folder, wait_for_download=wait_for_download)
    
    if not latest_file:
        print("   ❌ No Excel files found in downloads folder")
        print(f"   📁 Folder: {downloads_folder}")
        return False
    
    print(f"   📄 Found file: {os.path.basename(latest_file)}")
    print(f"   📅 Modified: {datetime.fromtimestamp(os.path.getmtime(latest_file)).strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"   📍 Full path: {latest_file}")
    
    # Run macro
    success = run_dt_macro(latest_file)
    
    if success:
        print("\n" + "="*80)
        print("✅ FILE CONVERTED SUCCESSFULLY!")
        print("="*80)
        print(f"   📁 File: {os.path.basename(latest_file)}")
        print(f"   📍 Location: {latest_file}")
        print("="*80)
        return True
    else:
        print("\n" + "="*80)
        print("❌ FILE CONVERSION FAILED")
        print("="*80)
        print("   ⚠️  The file was downloaded but the macro could not be executed.")
        print("   💡 You may need to run the macro manually in Excel.")
        return False


def process_all_downloads(downloads_folder=None, wait_for_download=True):
    """
    Run the DT macro on every Excel file found in the downloads folder.
    """
    if downloads_folder is None:
        BASE_DIR = Path(__file__).resolve().parents[2]
        downloads_folder = BASE_DIR / "data" / "downloads"

    downloads_path = Path(downloads_folder)

    if not downloads_path.exists():
        print(f"   ❌ Downloads folder does not exist: {downloads_path}")
        return False

    excel_files = sorted(downloads_path.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)

    if not excel_files:
        print("   ❌ No Excel files found in downloads folder")
        return False

    overall_success = True

    shared_app = None
    try:
        import xlwings as xw  # type: ignore

        existing_apps = list(xw.apps)
        if existing_apps:
            shared_app = existing_apps[0]
            try:
                shared_app.visible = True
            except Exception:
                pass
            print("   🟢 Reusing existing Excel instance for macro processing.")
        else:
            shared_app = xw.App(visible=True)
            print("   🟢 Launched Excel for macro processing.")

        try:
            shared_app.api.DisplayAlerts = False
            shared_app.api.ScreenUpdating = True
        except Exception:
            pass
    except ImportError:
        print("   ℹ️  xlwings not available; will use Python fallback if needed.")
        shared_app = None
    except Exception as app_error:
        print(f"   ⚠️  Could not prepare shared Excel instance: {app_error}")
        shared_app = None

    print("\n" + "="*80)
    print(f"📂 Processing {len(excel_files)} downloaded file(s)")
    print("="*80)

    for file_path in excel_files:
        file_str = str(file_path)
        print(f"\n   📄 Processing: {file_path.name}")

        if wait_for_download:
            if wait_for_file_download(file_str, max_wait=60):
                print("      ✅ File download complete")
            else:
                print("      ⚠️  File may still be downloading; proceeding with caution...")

        success = run_dt_macro(file_str, app=shared_app)
        if not success:
            overall_success = False

    if shared_app is not None:
        print("\n   🪟 Excel window(s) remain open for review. Close manually when finished.")

    return overall_success


if __name__ == "__main__":
    process_downloaded_file()

